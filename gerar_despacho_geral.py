from __future__ import annotations

import html
import re
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
XLSX_PATH = Path(r"C:\Users\higoo\Downloads\Planilha sem título (13).xlsx")
TEMPLATE_PATH = Path(r"C:\Users\higoo\OneDrive\Desktop\Despacho_Geral.html")
OUTPUT_PATH = ROOT / "despacho_geral.html"


def html_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    escaped = html.escape(text, quote=True)
    return escaped.encode("ascii", "xmlcharrefreplace").decode("ascii")


def date_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def read_template(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def build_entries(ws) -> list[dict[str, str]]:
    entries = []
    for base_row in range(4, 285, 40):
        if str(ws[f"E{base_row}"].value or "").strip() == "-":
            continue

        start_row = base_row + 6
        for row in range(start_row, start_row + 11):
            al_of_pm = ws[f"D{row}"].value
            re_value = ws[f"E{row}"].value
            process_number = ws[f"Q{row}"].value
            if not any((al_of_pm, re_value, process_number)):
                continue
            entries.append(
                {
                    "al_of_pm": html_text(al_of_pm),
                    "re": html_text(re_value),
                    "process_number": html_text(process_number),
                }
            )
    return entries


def build_list(entries: list[dict[str, str]]) -> str:
    lines = []
    for index, entry in enumerate(entries, start=1):
        suffix = "" if index == len(entries) else ";"
        line = (
            f'{index})N&ordm; do processo {entry["process_number"]}, '
            f'Cad PM {entry["re"]} {entry["al_of_pm"]}{suffix}'
        )
        if index < len(entries):
            line += "<br />"
        lines.append(line)
    return "\n".join(lines)


def replace_process_list(template: str, process_list: str) -> str:
    pattern = re.compile(
        r'(<p[^>]*class="Texto_Alinhado_Esquerda_Det"[^>]*>)\s*'
        r"1\)N(?:&ordm;|º) do processo\s*A260016958[\s\S]*?"
        r"</p>\s*(?=<p style=\"margin-top:15px)",
        flags=re.IGNORECASE,
    )
    if not pattern.search(template):
        raise RuntimeError("Nao encontrei a lista modelo de processos no HTML.")
    return pattern.sub(lambda match: f"{match.group(1)}{process_list}</p>\n\n", template, count=1)


def trim_after_last_responsible(template: str) -> str:
    marker = "CAD PM Resp. Restr/Conval/LTS"
    marker_index = template.rfind(marker)
    if marker_index == -1:
        return template

    paragraph_end = template.find("</p>", marker_index)
    cut_index = marker_index + len(marker) if paragraph_end == -1 else paragraph_end + 4
    closing_match = re.search(r"</body>[\s\S]*?</html>\s*$", template, flags=re.IGNORECASE)
    closing = closing_match.group(0) if closing_match else "\n</body>\n</html>"
    return template[:cut_index] + "\n" + closing


def ensure_utf8_meta(template: str) -> str:
    if re.search(r"<meta[^>]+charset=", template, flags=re.IGNORECASE):
        return re.sub(
            r"<meta[^>]+charset=[\"']?[^\"'>\s]+[\"']?[^>]*>",
            '<meta charset="utf-8">',
            template,
            count=1,
            flags=re.IGNORECASE,
        )
    return re.sub(r"<head[^>]*>", lambda match: match.group(0) + "\n<meta charset=\"utf-8\">", template, count=1, flags=re.IGNORECASE)


def main() -> None:
    workbook = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = workbook.active

    html_output = read_template(TEMPLATE_PATH)
    html_output = html_output.replace("015/31/26", html_text(ws["F2"].value) or "015/31/26")
    html_output = html_output.replace("20/05/2026", html_text(date_text(ws["B1"].value)) or "20/05/2026")
    html_output = replace_process_list(html_output, build_list(build_entries(ws)))
    html_output = trim_after_last_responsible(html_output)
    html_output = ensure_utf8_meta(html_output)

    OUTPUT_PATH.write_text("\ufeff" + html_output, encoding="utf-8")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
